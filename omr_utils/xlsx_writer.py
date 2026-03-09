"""Write scoring summary to XLSX workbook with openpyxl."""

# PIP3 modules
import openpyxl
import openpyxl.styles


#============================================
def _build_summary_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
	graded_results: list) -> None:
	"""Populate the Summary sheet with student scores.

	Args:
		ws: worksheet to populate
		graded_results: list of graded dicts from grade_student
	"""
	# bold font for header row
	bold_font = openpyxl.styles.Font(bold=True)
	headers = [
		"Student ID", "Filename", "Number Blanks",
		"Number Multiple", "Raw Score", "Total", "Percentage",
	]
	for col_idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = bold_font
	# sort students by student_id
	sorted_results = sorted(graded_results, key=lambda g: g["student_id"])
	for row_idx, graded in enumerate(sorted_results, start=2):
		ws.cell(row=row_idx, column=1, value=graded["student_id"])
		ws.cell(row=row_idx, column=2, value=graded.get("filename", ""))
		ws.cell(row=row_idx, column=3, value=graded.get("num_blank", 0))
		ws.cell(row=row_idx, column=4, value=graded.get("num_multiple", 0))
		ws.cell(row=row_idx, column=5, value=graded["raw_score"])
		ws.cell(row=row_idx, column=6, value=graded["total_questions"])
		# format percentage to 1 decimal place
		pct_value = round(graded["percentage"], 1)
		ws.cell(row=row_idx, column=7, value=pct_value)


#============================================
def _build_detailed_grades_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
	graded_results: list, num_questions: int) -> None:
	"""Populate the Detailed Grades sheet with per-question scores.

	Args:
		ws: worksheet to populate
		graded_results: list of graded dicts
		num_questions: total number of questions
	"""
	bold_font = openpyxl.styles.Font(bold=True)
	green_fill = openpyxl.styles.PatternFill(
		start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
	)
	red_fill = openpyxl.styles.PatternFill(
		start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
	)
	# header row
	header_cell = ws.cell(row=1, column=1, value="Student ID")
	header_cell.font = bold_font
	fn_cell = ws.cell(row=1, column=2, value="Filename")
	fn_cell.font = bold_font
	for q in range(1, num_questions + 1):
		cell = ws.cell(row=1, column=q + 2, value=f"Q{q}")
		cell.font = bold_font
	# data rows sorted by student_id
	sorted_results = sorted(graded_results, key=lambda g: g["student_id"])
	for row_idx, graded in enumerate(sorted_results, start=2):
		ws.cell(row=row_idx, column=1, value=graded["student_id"])
		ws.cell(row=row_idx, column=2, value=graded.get("filename", ""))
		per_q = graded["per_question"]
		for q in range(1, num_questions + 1):
			score = per_q.get(q, -1)
			col = q + 2
			if score == 1:
				cell = ws.cell(row=row_idx, column=col, value=1)
				cell.fill = green_fill
			elif score == 0:
				cell = ws.cell(row=row_idx, column=col, value=0)
				cell.fill = red_fill
			# -1 means not graded, leave cell blank


#============================================
def _build_student_answers_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
	key_data: dict, student_results: list,
	num_questions: int) -> None:
	"""Populate the Student Answers sheet with answer letters.

	Args:
		ws: worksheet to populate
		key_data: answer key dict with student_id and answers
		student_results: list of student answer dicts
		num_questions: total number of questions
	"""
	bold_font = openpyxl.styles.Font(bold=True)
	# header row
	header_cell = ws.cell(row=1, column=1, value="Student ID")
	header_cell.font = bold_font
	fn_cell = ws.cell(row=1, column=2, value="Filename")
	fn_cell.font = bold_font
	for q in range(1, num_questions + 1):
		cell = ws.cell(row=1, column=q + 2, value=f"Q{q}")
		cell.font = bold_font
	# first data row is the answer key labeled "KEY"
	ws.cell(row=2, column=1, value="KEY")
	key_answers = key_data["answers"]
	for q in range(1, num_questions + 1):
		answer = key_answers.get(q, "")
		if answer:
			ws.cell(row=2, column=q + 2, value=answer)
	# student rows sorted by student_id
	sorted_students = sorted(
		student_results, key=lambda s: s["student_id"]
	)
	for row_idx, student in enumerate(sorted_students, start=3):
		ws.cell(row=row_idx, column=1, value=student["student_id"])
		ws.cell(row=row_idx, column=2, value=student.get("filename", ""))
		student_answers = student["answers"]
		for q in range(1, num_questions + 1):
			answer = student_answers.get(q, "")
			if answer:
				ws.cell(row=row_idx, column=q + 2, value=answer)


#============================================
def _build_question_analysis_sheet(ws: openpyxl.worksheet.worksheet.Worksheet,
	key_data: dict, graded_results: list,
	num_questions: int) -> None:
	"""Populate the Question Analysis sheet with per-question stats.

	Args:
		ws: worksheet to populate
		key_data: answer key dict
		graded_results: list of graded dicts
		num_questions: total number of questions
	"""
	bold_font = openpyxl.styles.Font(bold=True)
	yellow_fill = openpyxl.styles.PatternFill(
		start_color="FFFF00", end_color="FFFF00", fill_type="solid"
	)
	headers = [
		"Question", "Num Correct", "Num Incorrect",
		"Num Blank", "Pct Correct",
	]
	for col_idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = bold_font
	key_answers = key_data["answers"]
	# only include questions where the key has an answer
	row_idx = 2
	for q in range(1, num_questions + 1):
		key_answer = key_answers.get(q, "")
		if not key_answer:
			continue
		# count correct, incorrect, blank across all students
		num_correct = 0
		num_incorrect = 0
		num_blank = 0
		for graded in graded_results:
			status = graded.get("per_question_status", {}).get(q, "")
			score = graded["per_question"].get(q, -1)
			if status == "blank":
				num_blank += 1
			elif score == 1:
				num_correct += 1
			elif score == 0:
				num_incorrect += 1
			else:
				num_blank += 1
		total_graded = num_correct + num_incorrect + num_blank
		if total_graded > 0:
			pct_correct = round(num_correct / total_graded * 100, 1)
		else:
			pct_correct = 0.0
		ws.cell(row=row_idx, column=1, value=q)
		ws.cell(row=row_idx, column=2, value=num_correct)
		ws.cell(row=row_idx, column=3, value=num_incorrect)
		ws.cell(row=row_idx, column=4, value=num_blank)
		pct_cell = ws.cell(row=row_idx, column=5, value=pct_correct)
		# highlight questions below 50% correct
		if pct_correct < 50.0:
			pct_cell.fill = yellow_fill
		row_idx += 1


#============================================
def _build_blank_multiple_detail_sheet(
	ws: openpyxl.worksheet.worksheet.Worksheet,
	graded_results: list) -> None:
	"""Populate a per-student blank/multiple detail sheet."""
	bold_font = openpyxl.styles.Font(bold=True)
	headers = [
		"Student ID", "Filename", "Number Blanks", "Number Multiple",
		"Blank Questions", "Multiple Questions",
	]
	for col_idx, header in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx, value=header)
		cell.font = bold_font
	sorted_results = sorted(graded_results, key=lambda g: g["student_id"])
	for row_idx, graded in enumerate(sorted_results, start=2):
		blank_qs = graded.get("blank_questions", [])
		multiple_qs = graded.get("multiple_questions", [])
		ws.cell(row=row_idx, column=1, value=graded["student_id"])
		ws.cell(row=row_idx, column=2, value=graded.get("filename", ""))
		ws.cell(row=row_idx, column=3, value=graded.get("num_blank", 0))
		ws.cell(row=row_idx, column=4, value=graded.get("num_multiple", 0))
		ws.cell(row=row_idx, column=5, value=" ".join(
			f"q{q}" for q in blank_qs))
		ws.cell(row=row_idx, column=6, value=" ".join(
			f"q{q}" for q in multiple_qs))


#============================================
def write_scoring_summary(output_path: str, key_data: dict,
	student_results: list, graded_results: list) -> None:
	"""Write a multi-tab XLSX scoring summary workbook.

	Creates five sheets: Summary, Detailed Grades, Student Answers,
	Question Analysis, and Blank Multiple Detail.

	Args:
		output_path: path for the output .xlsx file
		key_data: answer key dict from csv_writer.read_answers_csv
		student_results: list of student answer dicts
		graded_results: list of graded dicts from grade_student
	"""
	wb = openpyxl.Workbook()
	# determine number of questions from key
	num_questions = max(key_data["answers"].keys()) if key_data["answers"] else 0
	# Tab 1: Summary (default sheet)
	ws_summary = wb.active
	ws_summary.title = "Summary"
	_build_summary_sheet(ws_summary, graded_results)
	# Tab 2: Detailed Grades
	ws_detail = wb.create_sheet("Detailed Grades")
	_build_detailed_grades_sheet(ws_detail, graded_results, num_questions)
	# Tab 3: Student Answers
	ws_answers = wb.create_sheet("Student Answers")
	_build_student_answers_sheet(
		ws_answers, key_data, student_results, num_questions
	)
	# Tab 4: Question Analysis
	ws_analysis = wb.create_sheet("Question Analysis")
	_build_question_analysis_sheet(
		ws_analysis, key_data, graded_results, num_questions
	)
	# Tab 5: Blank and multiple detail per student
	ws_blank_multi = wb.create_sheet("Blank Multiple Detail")
	_build_blank_multiple_detail_sheet(ws_blank_multi, graded_results)
	wb.save(output_path)
