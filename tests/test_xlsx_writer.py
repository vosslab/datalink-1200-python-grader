"""Tests for omr_utils.xlsx_writer using synthetic data."""

# Standard Library
import os

# PIP3 modules
import openpyxl
import pytest

# local repo modules
import omr_utils.xlsx_writer


#============================================
@pytest.fixture
def synthetic_data(tmp_path: os.PathLike) -> dict:
	"""Create synthetic key, student, and graded data for 2 students, 5 questions.

	Returns:
		dict with keys: xlsx_path, key_data, student_results, graded_results
	"""
	key_data = {
		"student_id": "KEY",
		"answers": {1: "A", 2: "B", 3: "C", 4: "D", 5: "E"},
		"confidences": {1: 0.5, 2: 0.5, 3: 0.5, 4: 0.5, 5: 0.5},
		"flags": "",
	}
	student1 = {
		"student_id": "S001",
		"filename": "scan_001",
		"answers": {1: "A", 2: "B", 3: "A", 4: "D", 5: ""},
		"confidences": {1: 0.4, 2: 0.3, 3: 0.2, 4: 0.5, 5: 0.0},
		"flags": "",
	}
	student2 = {
		"student_id": "S002",
		"filename": "scan_002",
		"answers": {1: "C", 2: "B", 3: "C", 4: "D", 5: "E"},
		"confidences": {1: 0.1, 2: 0.4, 3: 0.5, 4: 0.3, 5: 0.2},
		"flags": "",
	}
	student_results = [student1, student2]
	graded1 = {
		"student_id": "S001",
		"filename": "scan_001",
		"raw_score": 3,
		"total_questions": 5,
		"percentage": 60.0,
		"per_question": {1: 1, 2: 1, 3: 0, 4: 1, 5: -1},
		"low_confidence": [],
	}
	graded2 = {
		"student_id": "S002",
		"filename": "scan_002",
		"raw_score": 4,
		"total_questions": 5,
		"percentage": 80.0,
		"per_question": {1: 0, 2: 1, 3: 1, 4: 1, 5: 1},
		"low_confidence": [],
	}
	graded_results = [graded1, graded2]
	xlsx_path = str(tmp_path / "test_summary.xlsx")
	# generate the workbook
	omr_utils.xlsx_writer.write_scoring_summary(
		xlsx_path, key_data, student_results, graded_results
	)
	result = {
		"xlsx_path": xlsx_path,
		"key_data": key_data,
		"student_results": student_results,
		"graded_results": graded_results,
	}
	return result


#============================================
def test_output_file_created(synthetic_data: dict) -> None:
	"""Verify the XLSX file exists after write."""
	assert os.path.isfile(synthetic_data["xlsx_path"])


#============================================
def test_workbook_has_four_sheets(synthetic_data: dict) -> None:
	"""Verify workbook has 4 sheets with correct names."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	expected = ["Summary", "Detailed Grades", "Student Answers",
		"Question Analysis"]
	assert wb.sheetnames == expected


#============================================
def test_summary_row_count(synthetic_data: dict) -> None:
	"""Verify Summary sheet has header + 2 student rows."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	ws = wb["Summary"]
	# max_row should be 3 (1 header + 2 students)
	assert ws.max_row == 3


#============================================
def test_summary_values(synthetic_data: dict) -> None:
	"""Verify Summary sheet contains correct student scores."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	ws = wb["Summary"]
	# row 2 = S001 (sorted first), row 3 = S002
	assert ws.cell(row=2, column=1).value == "S001"
	assert ws.cell(row=2, column=2).value == "scan_001"
	assert ws.cell(row=2, column=3).value == 3
	assert ws.cell(row=2, column=4).value == 5
	assert ws.cell(row=2, column=5).value == 60.0
	assert ws.cell(row=3, column=1).value == "S002"
	assert ws.cell(row=3, column=2).value == "scan_002"
	assert ws.cell(row=3, column=3).value == 4
	assert ws.cell(row=3, column=4).value == 5
	assert ws.cell(row=3, column=5).value == 80.0


#============================================
def test_detailed_grades_values(synthetic_data: dict) -> None:
	"""Verify Detailed Grades sheet has correct 1/0/None cells."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	ws = wb["Detailed Grades"]
	# row 2 = S001: per_question {1:1, 2:1, 3:0, 4:1, 5:-1}
	assert ws.cell(row=2, column=2).value == "scan_001"  # Filename
	assert ws.cell(row=2, column=3).value == 1   # Q1 correct
	assert ws.cell(row=2, column=4).value == 1   # Q2 correct
	assert ws.cell(row=2, column=5).value == 0   # Q3 incorrect
	assert ws.cell(row=2, column=6).value == 1   # Q4 correct
	assert ws.cell(row=2, column=7).value is None  # Q5 not graded
	# row 3 = S002: per_question {1:0, 2:1, 3:1, 4:1, 5:1}
	assert ws.cell(row=3, column=3).value == 0   # Q1 incorrect
	assert ws.cell(row=3, column=4).value == 1   # Q2 correct
	assert ws.cell(row=3, column=7).value == 1   # Q5 correct


#============================================
def test_student_answers_values(synthetic_data: dict) -> None:
	"""Verify Student Answers sheet has KEY row and answer letters."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	ws = wb["Student Answers"]
	# row 2 should be the KEY row
	assert ws.cell(row=2, column=1).value == "KEY"
	assert ws.cell(row=2, column=3).value == "A"  # Q1
	assert ws.cell(row=2, column=4).value == "B"  # Q2
	assert ws.cell(row=2, column=7).value == "E"  # Q5
	# row 3 = S001 answers
	assert ws.cell(row=3, column=1).value == "S001"
	assert ws.cell(row=3, column=2).value == "scan_001"  # Filename
	assert ws.cell(row=3, column=3).value == "A"  # Q1
	assert ws.cell(row=3, column=5).value == "A"  # Q3 (wrong answer)
	assert ws.cell(row=3, column=7).value is None  # Q5 blank
	# row 4 = S002 answers
	assert ws.cell(row=4, column=1).value == "S002"
	assert ws.cell(row=4, column=3).value == "C"  # Q1


#============================================
def test_question_analysis_counts(synthetic_data: dict) -> None:
	"""Verify Question Analysis sheet has correct counts."""
	wb = openpyxl.load_workbook(synthetic_data["xlsx_path"])
	ws = wb["Question Analysis"]
	# All 5 questions have key answers, so 5 data rows
	# Q1: S001=correct(1), S002=incorrect(0) -> 1 correct, 1 incorrect
	assert ws.cell(row=2, column=1).value == 1   # question number
	assert ws.cell(row=2, column=2).value == 1   # num correct
	assert ws.cell(row=2, column=3).value == 1   # num incorrect
	assert ws.cell(row=2, column=4).value == 0   # num blank
	assert ws.cell(row=2, column=5).value == 50.0  # pct correct
	# Q2: both correct -> 2 correct, 0 incorrect
	assert ws.cell(row=3, column=2).value == 2
	assert ws.cell(row=3, column=3).value == 0
	assert ws.cell(row=3, column=5).value == 100.0
	# Q5: S001=-1 (blank), S002=1 (correct) -> 1 correct, 0 incorrect, 1 blank
	assert ws.cell(row=6, column=2).value == 1   # num correct
	assert ws.cell(row=6, column=3).value == 0   # num incorrect
	assert ws.cell(row=6, column=4).value == 1   # num blank
